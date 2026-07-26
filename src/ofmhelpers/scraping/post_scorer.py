"""
filter_reels.py
---------------
Loads an existing instagram_reels.xlsx, filters low-performing reels,
adds engagement rate columns + avg views/day + score, sorts by score,
and saves as a new timestamped copy.

Usage:
    python filter_reels.py
    python filter_reels.py --input my_reels.xlsx
"""

from __future__ import annotations

import re
import shutil
from datetime import UTC, date, datetime
from pathlib import Path
from typing import TYPE_CHECKING, ClassVar

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ofmhelpers.log import get_logger

logger = get_logger(__name__)
if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

    from ofmhelpers.config.scrapers import ScraperConfig

# ── Thresholds ────────────────────────────────────────────────────────────────

VIEWS_THRESHOLD_DEFAULT = 40_000
VIEWS_THRESHOLD_TODAY = 10_000

# ── Ranking weights ───────────────────────────────────────────────────────────

WEIGHTS = {
    "views": 0.40,
    "like_rate": 0.30,
    "comment_rate": 0.20,
    "velocity": 0.10,
}


# ── Helpers ───────────────────────────────────────────────────────────────────

XML_BAD = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def _int(v):
    try:
        return int(v) if v is not None else None
    except (ValueError, TypeError):
        return None


def _float(v):
    try:
        return float(v) if v is not None else None
    except (ValueError, TypeError):
        return None


def _safe(v) -> str:
    return XML_BAD.sub("", str(v)) if v is not None else ""


def _parse_date(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        # Scraped timestamps carry no UTC offset; treat them as UTC so the
        # age comparison against `date.today()` below has a single basis.
        return (
            datetime.strptime(str(v)[:16], "%Y-%m-%d %H:%M").replace(tzinfo=UTC).date()
        )
    except Exception:
        return None


def _days_since(posted: date | None) -> float | None:
    if posted is None:
        return None
    return max((datetime.now(UTC).date() - posted).days, 1)


def _normalize(values: list) -> list[float]:
    clean = [v for v in values if v is not None]
    if not clean or max(clean) == min(clean):
        return [0.0] * len(values)
    lo, hi = min(clean), max(clean)
    return [((v - lo) / (hi - lo)) * 100 if v is not None else 0.0 for v in values]


# ── openpyxl style copy helpers ───────────────────────────────────────────────


def copy_font(f):
    return Font(
        name=f.name,
        size=f.size,
        bold=f.bold,
        italic=f.italic,
        color=f.color.rgb if f.color and f.color.type == "rgb" else "000000",
        underline=f.underline,
    )


def copy_fill(f):
    if f.fill_type and f.fill_type != "none":
        try:
            return PatternFill("solid", start_color=f.fgColor.rgb)
        except (AttributeError, ValueError, TypeError):
            # theme/indexed fills have no usable .rgb -- fall back to blank
            return PatternFill()
    return PatternFill()


def copy_alignment(a):
    return Alignment(
        horizontal=a.horizontal, vertical=a.vertical, wrap_text=a.wrap_text
    )


# ── Core processor ────────────────────────────────────────────────────────────


class PostFilterProcessor:
    NEW_HEADERS: ClassVar[list[str]] = [
        "Like Rate (%)",
        "Comment Rate (%)",
        "Avg Views/Day",
        "Score",
    ]

    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    HEADER_FILL = PatternFill("solid", start_color="1a1a2e")
    ALT_FILL = PatternFill("solid", start_color="F2F2F7")
    BODY_FONT = Font(name="Arial", size=10)
    NEW_HEADER_FILL = PatternFill("solid", start_color="0d3b6e")
    SCORE_HEADER_FILL = PatternFill("solid", start_color="1a4731")

    # Below this score a row is tinted flat grey instead of green-scaled.
    SCORE_HIGHLIGHT_MIN = 30

    # Row 1 is the header, so data starts at row 2.
    FIRST_DATA_ROW = 2

    def process(self, src_path: str, sheet_configs: dict[str, ScraperConfig]) -> str:
        """
        sheet_configs: maps sheet name -> ScraperConfig
        Sheets not in the dict are skipped.
        """
        src = Path(src_path)
        ts = datetime.now(UTC).strftime("%Y%m%d_%H_%M_%S")
        dst = src.parent / f"{src.stem}_filtered_{ts}.xlsx"

        shutil.copy2(src, dst)
        logger.info("copied %s -> %s", src.name, dst.name)

        wb = load_workbook(dst)
        for sheet_name in wb.sheetnames:
            if sheet_name not in sheet_configs:
                logger.info("skipping %r -- no config provided", sheet_name)
                continue
            self._process_sheet(wb[sheet_name], sheet_configs[sheet_name])

        wb.save(dst)
        logger.info("saved %s", dst)
        return str(dst)

    def _process_sheet(self, ws: Worksheet, config: ScraperConfig) -> None:
        if ws.max_row < self.FIRST_DATA_ROW:
            return  # header only, no data rows

        header = [cell.value for cell in ws[1]]
        try:
            col_views = header.index("Views")
            col_likes = header.index("Likes")
            col_comments = header.index("Comments")
            col_date = header.index("Date Posted")
        except ValueError:
            logger.warning("skipping %r -- missing expected columns", ws.title)
            return

        # ── Read all data rows into memory ────────────────────────────────────
        all_rows: list[list[dict]] = []
        for row in range(self.FIRST_DATA_ROW, ws.max_row + 1):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                row_data.append(
                    {
                        "value": cell.value,
                        "font": copy_font(cell.font),
                        "fill": copy_fill(cell.fill),
                        "alignment": copy_alignment(cell.alignment),
                        "number_format": cell.number_format,
                    }
                )
            all_rows.append(row_data)

        # ── Filter & compute metrics ──────────────────────────────────────────
        today = datetime.now(UTC).date()
        kept_rows: list[list[dict]] = []
        metrics: list[dict] = []
        removed = 0

        for row_data in all_rows:
            if not any(cd["value"] for cd in row_data):
                removed += 1
                continue

            views = _int(row_data[col_views]["value"])
            likes = _int(row_data[col_likes]["value"])
            comments = _int(row_data[col_comments]["value"])
            posted = _parse_date(row_data[col_date]["value"])

            threshold = (
                config.views_threshold_today
                if posted == today
                else config.views_threshold_default
            )
            if views is None or views < threshold:
                removed += 1
                continue

            days = _days_since(posted)
            like_rate = round(likes / views, 4) if views and likes else None
            comm_rate = round(comments / views, 4) if views and comments else None
            avg_vpd = round(views / days) if views and days else None

            kept_rows.append(row_data)
            metrics.append(
                {
                    "views": views,
                    "like_rate": like_rate,
                    "comm_rate": comm_rate,
                    "avg_vpd": avg_vpd,
                }
            )

        if not kept_rows:
            logger.info("sheet %r: 0 rows after filtering", ws.title)
            for row in range(ws.max_row, 1, -1):
                ws.delete_rows(row)
            return

        # ── Normalize & score using config weights ────────────────────────────
        w = config.content_ranking_weights
        norm_views = _normalize([m["views"] for m in metrics])
        norm_like = _normalize([m["like_rate"] for m in metrics])
        norm_comment = _normalize([m["comm_rate"] for m in metrics])
        norm_vel = _normalize([m["avg_vpd"] for m in metrics])

        scores = [
            round(
                norm_views[i] * w.views
                + norm_like[i] * w.like_rate
                + norm_comment[i] * w.comment_rate
                + norm_vel[i] * w.velocity,
                2,
            )
            for i in range(len(metrics))
        ]

        # ── Attach computed columns ───────────────────────────────────────────
        for i, row_data in enumerate(kept_rows):
            m, score = metrics[i], scores[i]

            row_data.append(
                {
                    "value": m["like_rate"],
                    "font": Font(name="Arial", size=10),
                    "fill": PatternFill(),
                    "alignment": Alignment(horizontal="center"),
                    "number_format": "0.00%",
                }
            )
            row_data.append(
                {
                    "value": m["comm_rate"],
                    "font": Font(name="Arial", size=10),
                    "fill": PatternFill(),
                    "alignment": Alignment(horizontal="center"),
                    "number_format": "0.00%",
                }
            )
            row_data.append(
                {
                    "value": m["avg_vpd"],
                    "font": Font(name="Arial", size=10),
                    "fill": PatternFill(),
                    "alignment": Alignment(horizontal="center"),
                    "number_format": "#,##0",
                }
            )

            intensity = int(255 - (score / 100) * 180)
            score_fill = (
                PatternFill("solid", start_color=f"00{intensity:02X}00")
                if score > self.SCORE_HIGHLIGHT_MIN
                else PatternFill("solid", start_color="CCCCCC")
            )
            row_data.append(
                {
                    "value": score,
                    "font": Font(name="Arial", size=10, bold=True),
                    "fill": score_fill,
                    "alignment": Alignment(horizontal="center"),
                    "number_format": "General",
                }
            )

        kept_rows.sort(key=lambda r: r[-1]["value"] or 0, reverse=True)

        # ── Rewrite sheet ─────────────────────────────────────────────────────
        for row in range(ws.max_row, 1, -1):
            ws.delete_rows(row)

        orig_col_count = len(header)
        for offset, label in enumerate(self.NEW_HEADERS):
            col_idx = orig_col_count + offset + 1
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            cell.fill = (
                self.SCORE_HEADER_FILL if label == "Score" else self.NEW_HEADER_FILL
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = 16

        for i, row_data in enumerate(kept_rows):
            dest_row = i + 2
            is_alt = dest_row % 2 == 0
            for col_idx, cd in enumerate(row_data, start=1):
                cell = ws.cell(row=dest_row, column=col_idx, value=cd["value"])
                cell.font = cd["font"]
                cell.alignment = cd["alignment"]
                cell.number_format = cd["number_format"]
                is_score_col = col_idx == len(row_data)
                cell.fill = (
                    cd["fill"]
                    if is_score_col
                    else (self.ALT_FILL if is_alt else PatternFill())
                )

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        logger.info(
            "sheet %r: kept %d, removed %d low-view rows",
            ws.title,
            len(kept_rows),
            removed,
        )
        logger.info(
            "sheet %r top row: %s | score %s",
            ws.title,
            kept_rows[0][0]["value"],
            kept_rows[0][-1]["value"],
        )
