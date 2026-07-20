from __future__ import annotations

import csv
import logging
import re
from typing import Optional
from ofmhelpers.scraping.models import PostBase
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class PostExcelExporter:
    HEADER = [
        "Username",
        "URL",
        "Date Posted",
        "Views",
        "Likes",
        "Comments",
        "Duration (s)",
        "Caption",
        "Hashtags",
    ]
    COL_WIDTHS = [18, 45, 20, 12, 12, 12, 14, 60, 50]

    HEADER_FILL = PatternFill("solid", start_color="1a1a2e")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    ALT_FILL = PatternFill("solid", start_color="F2F2F7")
    BODY_FONT = Font(name="Arial", size=10)
    LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")

    _SHEET_BAD = re.compile(r"[:\\/?*\[\]]")
    _XML_BAD = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")

    # ── sanitisers ────────────────────────────────────────────────────────────

    def _clean_sheet_name(self, raw: str) -> str:
        return self._SHEET_BAD.sub("_", raw).strip()[:30] or "sheet"

    def _to_str(self, v) -> str:
        return self._XML_BAD.sub("", str(v)) if v is not None else ""

    def _to_int(self, v) -> Optional[int]:
        try:
            return int(v) if v is not None else None
        except (ValueError, TypeError):
            return None

    def _to_float(self, v) -> Optional[float]:
        try:
            return float(v) if v is not None else None
        except (ValueError, TypeError):
            return None

    # ── public API ────────────────────────────────────────────────────────────

    def export(
        self, sheets: list[tuple[str, list[PostBase]]], output_path: str
    ) -> None:
        wb = Workbook()
        first = True

        for sheet_name, posts in sheets:
            ws = wb.active if first else wb.create_sheet()
            ws.title = self._clean_sheet_name(sheet_name)
            self._write_sheet(ws, posts)
            first = False

        try:
            wb.save(output_path)
            print(f"\nSaved -> {output_path}")
        except Exception as exc:
            all_posts = [post for _, posts in sheets for post in posts]
            self._fallback_csv(all_posts, output_path, exc)

    # ── internals ─────────────────────────────────────────────────────────────

    def _post_row(self, post: PostBase) -> list:
        return [
            self._to_str(post.username),
            self._to_str(post.url),
            post.timestamp.strftime("%Y-%m-%d %H:%M"),
            self._to_int(post.views),
            self._to_int(post.likes),
            self._to_int(post.comments),
            self._to_float(post.duration_seconds),
            self._to_str(post.caption),
            self._to_str(", ".join(f"#{h}" for h in post.hashtags)),
        ]

    def _write_sheet(self, ws, posts: list[PostBase]) -> None:
        ws.append(self.HEADER)

        for col_idx, (cell, width) in enumerate(zip(ws[1], self.COL_WIDTHS), start=1):
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 22

        for row_idx, post in enumerate(posts, start=2):
            try:
                ws.append(self._post_row(post))
            except Exception as exc:
                print(f"  Warning: skipped row {row_idx} ({post.username}) — {exc}")
                continue

            fill = self.ALT_FILL if row_idx % 2 == 0 else None
            for col_idx, cell in enumerate(ws[row_idx], start=1):
                cell.font = self.LINK_FONT if col_idx == 2 else self.BODY_FONT
                cell.alignment = Alignment(
                    vertical="center", wrap_text=col_idx in (8, 9)
                )
                if fill:
                    cell.fill = fill

            if post.url.startswith("http"):
                try:
                    ws.cell(row=row_idx, column=2).hyperlink = post.url
                except Exception:
                    pass

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _fallback_csv(
        self, posts: list[PostBase], xlsx_path: str, exc: Exception
    ) -> None:
        csv_path = xlsx_path.replace(".xlsx", ".csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(self.HEADER)
            writer.writerows(self._post_row(post) for post in posts)
        print(f"\nFailed to save xlsx ({exc}) — dumped to {csv_path}")
