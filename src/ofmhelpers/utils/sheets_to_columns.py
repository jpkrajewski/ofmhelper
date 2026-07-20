"""
sheets_columns_to_keys.py

Reads an uploaded profiles sheet (.xlsx) with one column per scraper
(e.g. INSTAGRAM_PROFILES, TIKTOK_PROFILES) and returns a dict keyed by
the Scrapers enum, each holding the non-empty values from that column.
"""

from pathlib import Path

import openpyxl

from ofmhelpers.config.scrapers import Scrapers


def sheets_columns_to_keys(path: str | Path) -> dict[Scrapers, list[str]]:
    """
    Expects a sheet where the header row contains column names matching
    Scrapers enum values exactly (e.g. "INSTAGRAM_PROFILES", "TIKTOK_PROFILES").
    Columns for scrapers not present in the sheet are returned as empty lists.
    Extra columns not matching any Scrapers value are ignored.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        raise ValueError(f"Sheet at {path} is empty — no header row found")

    # column name -> column index
    col_index: dict[str, int] = {
        str(name).strip(): idx for idx, name in enumerate(header) if name is not None
    }

    result: dict[Scrapers, list[str]] = {scraper: [] for scraper in Scrapers}

    for scraper in Scrapers:
        idx = col_index.get(scraper.value)
        if idx is None:
            continue  # column not present in this sheet, leave empty
        for row in ws.iter_rows(min_row=2, values_only=True):
            if idx >= len(row):
                continue
            value = row[idx]
            if value is None:
                continue
            value = str(value).strip()
            if value:
                result[scraper].append(value)

    wb.close()
    return result
